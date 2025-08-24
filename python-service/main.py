import logging
import os
import tempfile
from typing import Any, Dict, List, Optional

import openpyxl
import requests
from flask import Flask, jsonify, request
from flask_cors import CORS
from werkzeug.utils import secure_filename

app = Flask(__name__)

# Enable CORS for all routes
CORS(app, origins=['http://localhost:3000', 'http://localhost:5173', 'http://127.0.0.1:3000', 'http://127.0.0.1:5173'])

# Configuration
ALLOWED_EXTENSIONS = {"xlsx", "xls", "xlsm"}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
BACKEND_URL = os.environ.get("BACKEND_URL", "http://localhost:8000/api/mass-update")

# Logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def allowed_file(filename: str) -> bool:
    """Check if the uploaded file has an allowed extension"""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def validate_file_size(file) -> bool:
    """
    Check if file size is within limits.
    Works with Flask's FileStorage object (file.stream supports seek/tell).
    """
    try:
        file.stream.seek(0, os.SEEK_END)
        size = file.stream.tell()
    except Exception:
        # fallback to reading file content (less ideal)
        file.stream.seek(0)
        content = file.stream.read()
        size = len(content)
        # put pointer back
        file.stream.seek(0)
    finally:
        # reset pointer for subsequent operations
        try:
            file.stream.seek(0)
        except Exception:
            pass

    return size <= MAX_FILE_SIZE


def _normalize_header_value(value: Any) -> str:
    """Return a normalized header string (lowercase, stripped). Empty string for None."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip().lower()
    return str(value).strip().lower()


def process_excel_file(file_path: str) -> Dict[str, Any]:
    """Process Excel file and extract access card data."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        if ws is None:
            wb.close()
            return {"error": "Nav atrasts aktīvs darblapas Excel failā"}

        # Read header row (first row)
        header_cells = next(ws.iter_rows(min_row=1, max_row=1))
        headers: List[str] = [
            "" if cell.value is None else str(cell.value).strip()
            for cell in header_cells
        ]

        normalized_headers = [_normalize_header_value(h) for h in headers]
        header_indices = {h: idx for idx, h in enumerate(normalized_headers) if h != ""}

        # Required headers - QR Code is now optional
        required_headers = ["name", "group", "email"]
        missing_headers = [h for h in required_headers if h not in header_indices]

        if missing_headers:
            wb.close()
            return {
                "error": f'Trūkst nepieciešamo kolonnu: {", ".join(missing_headers)}'
            }

        # Check for optional QR Code column (various possible names)
        qr_code_variants = ["qr code", "qrcode", "qr_code", "qr-code", "uuid"]
        qr_code_index = None
        for variant in qr_code_variants:
            if variant in header_indices:
                qr_code_index = header_indices[variant]
                break

        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # skip completely empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            def get_row_value(index: Optional[int]) -> Optional[str]:
                if index is not None and index < len(row):
                    cell_value = row[index]
                    if cell_value is None:
                        return None
                    return (
                        str(cell_value).strip()
                        if not isinstance(cell_value, str)
                        else cell_value.strip()
                    )
                return None

            entry = {
                "name": get_row_value(header_indices.get("name")),
                "group": get_row_value(header_indices.get("group")),
                "email": get_row_value(header_indices.get("email")),
            }

            # Add QR code/UUID if column exists and has a value
            qr_code_value = get_row_value(qr_code_index)
            if qr_code_value and qr_code_value.strip():
                entry["qr_code"] = qr_code_value.strip()

            data.append(entry)

        wb.close()
        return {
            "success": True,
            "data": data,
            "total_records": len(data),
            "headers": headers,
            "has_qr_code_column": qr_code_index is not None,
        }

    except Exception as e:
        # ensure workbook closed in case of some exceptions
        try:
            wb.close()
        except Exception:
            pass
        return {"error": f"Kļūda apstrādājot Excel failu: {str(e)}"}


@app.route("/upload/excel", methods=["POST", "OPTIONS"])
def upload_excel():
    """Handle Excel file upload and send processed data to another backend"""
    # Handle preflight OPTIONS request
    if request.method == "OPTIONS":
        return jsonify({"status": "ok"}), 200

    temp_path = None
    try:
        if "file" not in request.files:
            return (
                jsonify({"status": "error", "message": "Nav izvēlēts fails"}),
                400,
            )

        file = request.files["file"]

        if not file or file.filename is None or file.filename == "":
            return (
                jsonify({"status": "error", "message": "Nav izvēlēts fails"}),
                400,
            )

        if not allowed_file(file.filename):
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Neatbalstīts faila formāts. Lūdzu augšupielādējiet .xlsx vai .xls failu",
                    }
                ),
                400,
            )

        if not validate_file_size(file):
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Fails ir par lielu. Maksimālais izmērs ir 10MB",
                    }
                ),
                400,
            )

        filename = secure_filename(file.filename)
        # Save to a temporary file
        with tempfile.NamedTemporaryFile(
                delete=False, suffix=f"-{filename}"
        ) as tmp_file:
            temp_path = tmp_file.name
            file.save(temp_path)

        # Process file
        result = process_excel_file(temp_path)

        # Clean up tmp file as soon as possible
        try:
            if temp_path and os.path.exists(temp_path):
                os.unlink(temp_path)
                temp_path = None
        except Exception as e:
            logger.warning("Neizdevās izdzēst pagaidu failu: %s", e)

        if "error" in result:
            return jsonify({"status": "error", "message": result["error"]}), 400

        # Send to backend
        try:
            payload = {
                "data": result["data"],
                "total_records": result["total_records"]
            }

            # Add employee_id if provided
            employee_id = request.form.get('employee_id')
            if employee_id:
                payload["employee_id"] = employee_id
                logger.info(
                    "Sūtu %d ierakstus uz backend %s ar darbinieku ID: %s",
                    result["total_records"], BACKEND_URL, employee_id
                )
            else:
                logger.info(
                    "Sūtu %d ierakstus uz backend %s", result["total_records"], BACKEND_URL
                )

            # Log QR code column detection
            if result.get("has_qr_code_column"):
                logger.info("QR Code kolonna atrasta Excel failā")

            response = requests.post(BACKEND_URL, json=payload, timeout=10)
        except requests.RequestException as e:
            logger.exception("Neizdevās izveidot savienojumu ar backend")
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": f"Nevar izveidot savienojumu ar backend: {str(e)}",
                    }
                ),
                502,
            )

        # Always return something based on backend response
        try:
            resp_text = response.text
        except Exception:
            resp_text = "<nepieejams atbildes saturs>"

        if response.status_code == 200:
            success_message = "Veiksmīgi apstrādāti dati un nosūtīti uz serveri"
            if result.get("has_qr_code_column"):
                success_message += " (ar QR kodu kolonu)"

            return (
                jsonify(
                    {
                        "status": "success",
                        "message": success_message,
                        "total_records": result["total_records"],
                        "has_qr_code_column": result.get("has_qr_code_column", False),
                    }
                ),
                200,
            )
        else:
            # Forward backend status and a short message for debugging
            logger.warning(
                "Backend atbildēja ar %s: %s", response.status_code, resp_text
            )
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": f"Backend returned status {response.status_code}",
                        "backend_body": resp_text,
                    }
                ),
                502,
            )

    except Exception as e:
        logger.exception("Servera kļūda pie augšupielādes")
        return jsonify({"status": "error", "message": f"Servera kļūda: {str(e)}"}), 500
    finally:
        # final cleanup safeguard
        try:
            if temp_path and os.path.exists(temp_path):
                os.unlink(temp_path)
        except Exception:
            pass


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0")